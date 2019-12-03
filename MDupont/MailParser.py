import mailparser
from email.parser import BytesParser
from email_reply_parser import EmailReplyParser
from email import policy
import glob
import email
import string
from collections import Counter
import time
import openpyxl
import os
import csv



def tryWithDefault(cell, value, default):
    try:
        cell.value = value
    except Exception as e:
        cell.value = default

def grabAddress(inputTuple):
    (_, recipientAddress) = inputTuple
    return recipientAddress

def removeDuplicates(edgeTuple):
    (fromElem, toElem, emailValue) = edgeTuple
    return fromElem != toElem



start = time.time()

allEmails = glob.glob('emails/*.eml')
#useful for debugging without running through the entire email set.
# allEmails = allEmails[0:20]

#####################################################################################################################################
#                              Initializing Excel workbook to store output in multiple sheets. 
#####################################################################################################################################
outputExcelWorkbook = openpyxl.Workbook()
edgeSheet = outputExcelWorkbook.create_sheet("EdgeList")
attributeSheet = outputExcelWorkbook.create_sheet("Attributes")
unreadableEmails = outputExcelWorkbook.create_sheet("UnreadableEmails")

edgeSheet.title = "EdgeList"
attributeSheet.title = "Attributes"
unreadableEmails.title = "UnreadableEmails"



print(str(len(allEmails)) + " files found.") 
emailNetworkEdgeList = []                                                                           #type should be (str * str * str)      
unreadableEmailsList = []                                                                           #Useful metadata for us.
emailAttributesList = []                                                                            #Not used yet, might be useful for timestamp information



#####################################################################################################################################
#                                    Parsing each email into To, From, Body, and Date
#####################################################################################################################################
for thisFile in allEmails:
    
    email = os.path.basename(thisFile)                                                              #Removes path information
    
    #Extracting the head element of email body
    with open(thisFile, 'rb') as fp:
        msg = BytesParser(policy=policy.default).parse(fp)
    try:
        body = msg.get_body(preferencelist='plain').get_content()
        #Grabbing only the head element of the email reply list.
        body_msg = EmailReplyParser.read(body)
        if len(body_msg.fragments) > 0:
            body = body_msg.fragments[0].content
            body = body.replace("\r", "")
            body = body.replace("\n", " ")
        else:
            body = ""
    except Exception as e:
        body = ""
    mail = mailparser.parse_from_file(thisFile)

    # Handle From Address
    fromList = mail.from_
    if  len(fromList) == 0:
        outputLine = " had 0 \"From\" addresses - this email is likely unreadable"
        print(email + outputLine)
        unreadableEmailsList.append(email + "\n")

    elif len(fromList) > 1:
        outputLine = " had >1 \"From\" addresses - this email is likely unreadable"
        print(email + outputLine)
        unreadableEmailsList.append(email + outputLine + "\n")
    
    else:
        (_, senderAddress) = fromList[0]
        senderAddress = senderAddress.lower()
        recipientAddresses = map(grabAddress, mail.to)
        emailTime = mail.date

        #Write all To-From Edges
        for recipientAddress in recipientAddresses:
            edge = (senderAddress, recipientAddress, email)
            emailNetworkEdgeList.append(edge)
        emailAttributesList.append((email, emailTime, body))
        print("Parsed " + email + " successfully.")

#Remove self-referential emails
#Maybe ToDo - record which duplicate references were removed?
preSelfReferenceRemoval = len(emailNetworkEdgeList)
emailNetworkEdgeList = list(filter(removeDuplicates, emailNetworkEdgeList))
postSelfReferenceRemoval = len(list(emailNetworkEdgeList))
print("The number of removed self-referential edges is: " 
    + str(preSelfReferenceRemoval - postSelfReferenceRemoval))



#####################################################################################################################################
#                                           Write all email edges to xlsx file
#####################################################################################################################################
emailNetworkEdgeList.insert(0, ("from", "to", "email"))
for index in range(len(emailNetworkEdgeList)):
    row = index+ 1
    edge = emailNetworkEdgeList[index]
    (sender, recipient, email) = edge
    
    senderCell = edgeSheet.cell(row=row, column=1)
    recipientCell = edgeSheet.cell(row=row, column=2)
    emailCell = edgeSheet.cell(row=row, column=3)
    tryWithDefault(senderCell, sender, "")
    tryWithDefault(recipientCell, recipient, "")
    tryWithDefault(emailCell, email, "")



#####################################################################################################################################
#                                           Write all email attributes to xlsx file
#####################################################################################################################################
emailAttributesList.insert(0, ("email", "timestamp", "body"))
for index in range(len(emailAttributesList)):
    row = index + 1
    attributes = emailAttributesList[index]
    (email, timestamp, body) = attributes
    
    emailCell = attributeSheet.cell(row=row, column=1)
    timestampCell = attributeSheet.cell(row=row, column=2)
    bodyCell = attributeSheet.cell(row=row, column=3)
    tryWithDefault(emailCell, email, "")
    tryWithDefault(timestampCell, timestamp, "")
    tryWithDefault(bodyCell, body, "")



#####################################################################################################################################
#                             Write all emails we could not parse due to problems with addresses
#####################################################################################################################################
unreadableEmailsList.insert(0, "Unreadable Emails")
for index in range(len(unreadableEmailsList)):
    row = index + 1
    emailCell = attributeSheet.cell(row=row, column=1)
    tryWithDefault(emailCell, email, "")

#Save File.
outputExcelWorkbook.save("EmailDataset.xlsx")

print("Time elapsed(in seconds): " + str(time.time()-start))